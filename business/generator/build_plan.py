#!/usr/bin/env python3
"""Builds Lumenwright-Business-Plan.xlsx — 10 tabs, live forecast formulas."""
import content_a as A
import content_b as B
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

AMBER = "F5A623"; DARK = "1A1611"; PAPER = "FBF7EF"; INK = "2B2416"
DIM = "6B6353"; YELLOW = "FFF3C4"; GREEN = "E8F3EC"; LINE = "D8CFBE"

f_title = Font(name="Georgia", size=18, bold=True, color=INK)
f_h = Font(name="Georgia", size=13, bold=True, color=INK)
f_th = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
f_td = Font(name="Calibri", size=10, color=INK)
f_note = Font(name="Calibri", size=10, italic=True, color=DIM)
f_k = Font(name="Calibri", size=10, bold=True, color=INK)

fill_th = PatternFill("solid", fgColor=INK)
fill_h = PatternFill("solid", fgColor="EFE7D7")
fill_in = PatternFill("solid", fgColor=YELLOW)
fill_alt = PatternFill("solid", fgColor="F5F0E6")
fill_total = PatternFill("solid", fgColor=GREEN)
thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")

def write_sheet(wb, name, rows, widths):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    r = 1
    alt = False
    for style, vals in rows:
        if style == "b":
            r += 1; alt = False; continue
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = wrap
            if style == "t":
                cell.font = f_title
            elif style == "h":
                cell.font = f_h; cell.fill = fill_h
            elif style == "th":
                cell.font = f_th; cell.fill = fill_th; cell.border = border
            elif style == "n":
                cell.font = f_note
            elif style == "k":
                cell.font = f_k if c == 1 else f_td
                cell.border = border
            else:  # r
                cell.font = f_td; cell.border = border
                if alt: cell.fill = fill_alt
                if isinstance(v, (int, float)) and abs(v) > 500:
                    cell.number_format = '#,##0'
        if style == "h":
            # extend header fill across width
            for c in range(len(vals) + 1, len(widths) + 1):
                ws.cell(row=r, column=c).fill = fill_h
        if style == "r":
            alt = not alt
        else:
            alt = False
        if style == "t":
            ws.row_dimensions[r].height = 26
        r += 1
    ws.freeze_panes = "A2"
    return ws

def build_forecast(wb):
    ws = wb.create_sheet("5 · Forecast")
    ws.sheet_view.showGridLines = False
    N = 24  # months
    first_col = 3  # C
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 13
    for i in range(N + 1):
        ws.column_dimensions[get_column_letter(first_col + i)].width = 11

    months = []
    names = ["Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb","Mar","Apr","May","Jun"]
    yr = 26
    for i in range(N):
        m = names[i % 12]
        if m == "Jan": yr += 0  # label handles year below
        months.append(f"{m} '{26 + ((6 + i) // 12)}")

    def C(i):  # month column letter, i = 0-based month
        return get_column_letter(first_col + i)

    r = 1
    ws.cell(row=r, column=1, value="5 · 36-MONTH FORECAST — EDIT THE YELLOW CELLS; EVERYTHING RECALCULATES").font = f_title
    r += 1
    ws.cell(row=r, column=1, value="Base case. Yellow = your assumptions (tickets, margins, volumes, opex, funding). Formulas do the rest. Y3 column = Year-2 total × growth factor.").font = f_note
    r += 2

    # ── assumptions ──
    ws.cell(row=r, column=1, value="ASSUMPTIONS — UNIT ECONOMICS").font = f_h
    ws.cell(row=r, column=1).fill = fill_h
    r += 1
    assumptions = [
        ("Avg retrofit project ($)", 22000, "Avg GM %", 0.42),
        ("Avg restoration ($)", 2800, "Avg GM %", 0.60),
        ("Avg custom fixture ($)", 7500, "Avg GM %", 0.45),
        ("Avg Vela order ($)", 1800, "Avg GM %", 0.50),
        ("Service plan ($/contract/mo)", 49, "Avg GM %", 0.70),
        ("Avg design/audit fee ($)", 300, "Avg GM %", 0.90),
        ("Year-3 growth factor (× Y2)", 1.6, "", None),
    ]
    arow = {}
    for i, (label, val, l2, v2) in enumerate(assumptions):
        ws.cell(row=r, column=1, value=label).font = f_td
        cv = ws.cell(row=r, column=2, value=val); cv.fill = fill_in; cv.font = f_k
        cv.number_format = '#,##0' if val > 100 else '0.00'
        if v2 is not None:
            ws.cell(row=r, column=4, value=l2).font = f_td
            c2 = ws.cell(row=r, column=5, value=v2); c2.fill = fill_in; c2.font = f_k
            c2.number_format = '0%'
        arow[i] = r
        r += 1
    r += 1

    # header row of months
    hdr = r
    ws.cell(row=r, column=1, value="MONTHLY MODEL").font = f_th
    ws.cell(row=r, column=1).fill = fill_th
    ws.cell(row=r, column=2, value="Y1 total").font = f_th
    ws.cell(row=r, column=2).fill = fill_th
    for i, m in enumerate(months):
        c = ws.cell(row=r, column=first_col + i, value=m)
        c.font = f_th; c.fill = fill_th
    cy3 = ws.cell(row=r, column=first_col + N, value="Year 3")
    cy3.font = f_th; cy3.fill = fill_th
    ws.column_dimensions[get_column_letter(first_col + N)].width = 12
    r += 1

    def input_row(label, values, fmt='0'):
        nonlocal r
        ws.cell(row=r, column=1, value=label).font = f_td
        for i, v in enumerate(values):
            c = ws.cell(row=r, column=first_col + i, value=v)
            c.fill = fill_in; c.font = f_td; c.number_format = fmt; c.border = border
        row = r; r += 1
        return row

    def formula_row(label, fmla, fmt='#,##0', bold=False, total=True, y3=None):
        nonlocal r
        row = r
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = f_k if bold else f_td
        for i in range(N):
            c = ws.cell(row=r, column=first_col + i, value=fmla(i, row))
            c.number_format = fmt; c.border = border
            c.font = f_k if bold else f_td
            if bold: c.fill = fill_total
        if total:
            t = ws.cell(row=r, column=2, value=f"=SUM(C{r}:N{r})")
            t.number_format = fmt; t.font = f_k
        if y3:
            c = ws.cell(row=r, column=first_col + N, value=(y3(row) if callable(y3) else y3))
            c.number_format = fmt; c.font = f_k if bold else f_td; c.border = border
        r += 1
        return row

    # volume inputs (base-case ramp, all editable)
    v_retro = input_row("Retrofit projects closed", [0,0,0,1,1,0,1,1,1,1,1,1, 1,2,2,2,2,2,2,3,3,3,3,3])
    v_resto = input_row("Restorations completed",   [0,1,2,2,2,3,3,3,3,4,4,4, 4,4,5,5,5,5,6,6,6,6,7,7])
    v_cust  = input_row("Custom fixtures delivered",[0,0,0,0,1,1,1,1,1,1,1,1, 1,1,2,1,2,2,2,2,2,2,2,2])
    v_vela  = input_row("Vela orders",              [0,0,1,1,1,2,2,2,2,3,3,3, 3,4,4,4,5,5,5,6,6,6,7,7])
    v_svc   = input_row("NEW service contracts",    [0,0,0,1,1,0,1,1,1,1,1,1, 1,2,2,2,2,2,2,3,3,3,3,3])
    v_fees  = input_row("Design consults/audits (paid)", [2,3,4,5,6,6,7,7,8,8,8,8, 8,9,9,10,10,10,11,11,12,12,12,12])
    r += 1

    # cumulative service contracts
    svc_cum = formula_row(
        "Service contracts (cumulative)",
        lambda i, row: f"={C(i)}{v_svc}" if i == 0 else f"={get_column_letter(first_col+i-1)}{row}+{C(i)}{v_svc}",
        fmt='0', total=False,
        y3=lambda row: f"={C(N-1)}{row}*$B${arow[6]}")
    r += 1

    # revenue rows
    ws.cell(row=r, column=1, value="REVENUE").font = f_h; ws.cell(row=r, column=1).fill = fill_h; r += 1
    rev_rows = []
    streams = [
        ("Retrofit revenue", v_retro, arow[0]),
        ("Restoration revenue", v_resto, arow[1]),
        ("Custom fixture revenue", v_cust, arow[2]),
        ("Vela product revenue", v_vela, arow[3]),
    ]
    for label, vrow, a in streams:
        rev_rows.append(formula_row(label, lambda i, row, vr=vrow, aa=a: f"={C(i)}{vr}*$B${aa}",
                                    y3=lambda row: f"=SUM({C(12)}{row}:{C(23)}{row})*$B${arow[6]}"))
    rev_svc = formula_row("Service plan revenue",
                          lambda i, row: f"={C(i)}{svc_cum}*$B${arow[4]}",
                          y3=lambda row: f"=SUM({C(12)}{row}:{C(23)}{row})*$B${arow[6]}")
    rev_rows.append(rev_svc)
    rev_fees = formula_row("Design/audit fees",
                           lambda i, row: f"={C(i)}{v_fees}*$B${arow[5]}",
                           y3=lambda row: f"=SUM({C(12)}{row}:{C(23)}{row})*$B${arow[6]}")
    rev_rows.append(rev_fees)
    total_rev = formula_row("TOTAL REVENUE",
                            lambda i, row: "=" + "+".join(f"{C(i)}{rr}" for rr in rev_rows),
                            bold=True,
                            y3="=" + "+".join(f"{get_column_letter(first_col+N)}{rr}" for rr in rev_rows))
    r += 1

    # gross margin
    ws.cell(row=r, column=1, value="GROSS MARGIN").font = f_h; ws.cell(row=r, column=1).fill = fill_h; r += 1
    gm_rows = []
    gm_map = [(rev_rows[0], arow[0]), (rev_rows[1], arow[1]), (rev_rows[2], arow[2]),
              (rev_rows[3], arow[3]), (rev_svc, arow[4]), (rev_fees, arow[5])]
    labels = ["Retrofit GM", "Restoration GM", "Custom GM", "Vela GM", "Service GM", "Fees GM"]
    for (rr, a), lab in zip(gm_map, labels):
        gm_rows.append(formula_row(lab, lambda i, row, rr=rr, aa=a: f"={C(i)}{rr}*$E${aa}",
                                   y3=f"={get_column_letter(first_col+N)}{rr}*$E${a}"))
    total_gm = formula_row("TOTAL GROSS MARGIN",
                           lambda i, row: "=" + "+".join(f"{C(i)}{g}" for g in gm_rows), bold=True,
                           y3="=" + "+".join(f"{get_column_letter(first_col+N)}{g}" for g in gm_rows))
    gm_pct = formula_row("GM %", lambda i, row: f"=IF({C(i)}{total_rev}=0,0,{C(i)}{total_gm}/{C(i)}{total_rev})",
                         fmt='0%', total=False,
                         y3=f"={get_column_letter(first_col+N)}{total_gm}/{get_column_letter(first_col+N)}{total_rev}")
    ws.cell(row=gm_pct, column=2, value=f"=B{total_gm}/B{total_rev}").number_format = '0%'
    r += 1

    # opex inputs
    ws.cell(row=r, column=1, value="OPERATING EXPENSES (editable)").font = f_h; ws.cell(row=r, column=1).fill = fill_h; r += 1
    ox = []
    ox.append(input_row("Marketing", [500,500,500,2500,2500,2500,2500,3000,3000,4000,4000,4000, 4500,4500,5000,5000,5000,5500,5500,5500,6000,6000,6000,6000], '#,##0'))
    ox.append(input_row("Owner draw", [3500]*6 + [6000]*6 + [7000]*12, '#,##0'))
    ox.append(input_row("Staff payroll (non-COGS)", [0]*8 + [3500]*4 + [3500,3500,9500,9500,9500,9500,9500,14000,14000,14000,14000,14000], '#,##0'))
    ox.append(input_row("Facility + storage", [400]*8 + [2000]*16, '#,##0'))
    ox.append(input_row("Insurance", [900]*24, '#,##0'))
    ox.append(input_row("Vehicle (loan+fuel+maint)", [0,650,650,650,650,650,650,650,650,650,650,650] + [700]*12, '#,##0'))
    ox.append(input_row("Software/office/other", [750]*24, '#,##0'))
    ox.append(input_row("ONE-TIME setup costs", [12000,14000,14000,6000,3000,0,0,0,0,0,0,0] + [0]*12, '#,##0'))
    total_ox = formula_row("TOTAL OPEX", lambda i, row: "=" + "+".join(f"{C(i)}{o}" for o in ox), bold=True,
                           y3=lambda row: f"=SUM({C(12)}{row}:{C(23)}{row})*1.35")
    r += 1

    # bottom line
    ws.cell(row=r, column=1, value="CASH VIEW").font = f_h; ws.cell(row=r, column=1).fill = fill_h; r += 1
    ebitda = formula_row("Operating profit (GM − opex)",
                         lambda i, row: f"={C(i)}{total_gm}-{C(i)}{total_ox}", bold=True,
                         y3=f"={get_column_letter(first_col+N)}{total_gm}-{get_column_letter(first_col+N)}{total_ox}")
    fund = input_row("Funding inflows (owner/SBA/etc.)", [18500,0,50000,0,0,0,0,0,100000,0,0,0] + [0]*12, '#,##0')
    cash = formula_row("CUMULATIVE CASH",
                       lambda i, row: (f"={C(i)}{ebitda}+{C(i)}{fund}" if i == 0
                                  else f"={get_column_letter(first_col+i-1)}{row}+{C(i)}{ebitda}+{C(i)}{fund}"),
                       bold=True, total=False)
    r += 2

    ws.cell(row=r, column=1, value="SCENARIOS — HOW TO STRESS THIS MODEL").font = f_h; ws.cell(row=r, column=1).fill = fill_h; r += 1
    for note in [
        "Downside: cut every volume row 40% and delay the 7(a) row to $0 — cumulative cash must stay >$0. If not, cut the M9 hire and owner draw first (they're yellow for a reason).",
        "Upside: retrofit ramp +50% forces the journeyman hire 3 months earlier — check payroll row before celebrating.",
        "Rule: re-forecast on the 1st of every month with actuals; variance >20% on any stream two months running = strategy conversation, not a spreadsheet fix.",
    ]:
        ws.cell(row=r, column=1, value="• " + note).font = f_note
        r += 1

    ws.freeze_panes = "C%d" % (hdr + 1)
    return {"total_rev": total_rev, "cash": cash, "N": N}

wb = Workbook()
wb.remove(wb.active)

write_sheet(wb, "0 · Start Here", A.START_HERE, [26, 120])
write_sheet(wb, "1 · Business Plan", A.BUSINESS_PLAN, [30, 46, 40, 42, 20])
write_sheet(wb, "2 · GTM + Workback", A.GTM, [14, 24, 90, 14])
write_sheet(wb, "3 · Phased Approach", A.PHASED, [24, 18, 30, 60, 26, 60])
write_sheet(wb, "4 · Revenue Engine", A.REVENUE, [30, 16, 26, 16, 20, 30, 48])
build_forecast(wb)
write_sheet(wb, "6 · Designer Network", A.DESIGNER, [34, 60, 40, 30])
write_sheet(wb, "7 · Digital Marketing", B.MARKETING, [22, 34, 44, 26, 30, 24])
write_sheet(wb, "8 · Vendors + Sourcing", B.VENDORS, [26, 40, 70, 46])
write_sheet(wb, "9 · Funding + Startup", B.FUNDING, [26, 60, 34, 40, 26])

out = "Lumenwright-Business-Plan.xlsx"
wb.save(out)
print("saved", out)

# sanity: recompute base-case Y1 revenue in python to sync Start Here numbers
tickets = dict(retro=22000, resto=2800, cust=7500, vela=1800, svc=49, fee=300)
vols = dict(
    retro=[0,0,0,1,1,0,1,1,1,1,1,1], resto=[0,1,2,2,2,3,3,3,3,4,4,4],
    cust=[0,0,0,0,1,1,1,1,1,1,1,1], vela=[0,0,1,1,1,2,2,2,2,3,3,3],
    svc_new=[0,0,0,1,1,0,1,1,1,1,1,1], fee=[2,3,4,5,6,6,7,7,8,8,8,8])
cum = 0; svc_rev = 0
for n in vols["svc_new"]:
    cum += n; svc_rev += cum * tickets["svc"]
y1 = (sum(vols["retro"]) * tickets["retro"] + sum(vols["resto"]) * tickets["resto"]
      + sum(vols["cust"]) * tickets["cust"] + sum(vols["vela"]) * tickets["vela"]
      + svc_rev + sum(vols["fee"]) * tickets["fee"])
print("base-case Y1 revenue:", round(y1))
